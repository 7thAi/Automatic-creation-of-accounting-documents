"""
Модуль для заполнения адресного перечня (АП) данными из фотографий.
"""
import re
import logging
from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)


class APFiller:
    """Заполняет Excel-файл АП данными о фотографиях."""

    EXTENSIONS = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff'}
    ROW_START = 4

    FOLDER_MAP = {
        "АП ДТ": ("1. ДТ", "B", "C"),
        "АП МКД": ("2. МКД", "C", "D"),
        "АП ОДХ": ("3. ОДХ", "B", "C"),
        "АП ОО": ("4. ОО", "B", "C")
    }

    COUNT_CELLS = {
        "ДТ": ("АП ДТ", "G2"),
        "ДТ_пройденные": ("АП ДТ", "H2"),
        "МКД": ("АП МКД", "H2"),
        "МКД_пройденные": ("АП МКД", "I2"),
        "ОДХ": ("АП ОДХ", "G2"),
        "ОДХ_пройденные": ("АП ОДХ", "H2"),
        "ОО": ("АП ОО", "G2"),
        "ОО_пройденные": ("АП ОО", "H2")
    }

    def clean_name(self, name: str) -> str:
        """Очищает имя файла от номера в скобках и заменяет подчёркивания на слэши.
        
        Args:
            name: Исходное имя файла.
            
        Returns:
            Очищенное имя.
        """
        name_without_ext = Path(name).stem
        return re.sub(r"\s*\(\d+\)$", "", name_without_ext).replace("_", "/").strip()

    def get_all_files_with_subfolders(self, folder_path: Path) -> list:
        """Получает список всех файлов изображений с информацией о подпапках.
        
        Args:
            folder_path: Путь к папке для сканирования.
            
        Returns:
            Список словарей с информацией о файлах.
        """
        if not folder_path.exists():
            logger.warning(f"Папка не существует: {folder_path}")
            return []
            
        files_info = [
            {
                "filename": item.name,
                "subfolder": item.parent.name if item.parent != folder_path else folder_path.name
            }
            for item in sorted(folder_path.rglob("*"))
            if item.is_file() and item.suffix.lower() in self.EXTENSIONS
        ]
        return files_info

    def _style_cell(self, cell, bold: bool = False) -> None:
        """Применяет стиль к ячейке.
        
        Args:
            cell: Ячейка Excel для стилизации.
            bold: Использовать жирный шрифт.
        """
        thin = Side(border_style="thin", color="000000")
        cell.font = Font(name="Times New Roman", size=14, bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def fill_counts(self, excel_path: Path, counts: Dict[str, int]) -> None:
        """Заполняет ячейки с количеством объектов.
        
        Args:
            excel_path: Путь к файлу Excel.
            counts: Словарь с количеством объектов по категориям.
        """
        wb = None
        try:
            wb = load_workbook(str(excel_path), keep_vba=True)
            for key, value in counts.items():
                if key in self.COUNT_CELLS:
                    sheet_name, cell_ref = self.COUNT_CELLS[key]
                    if sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        sheet[cell_ref] = int(value)
                        self._style_cell(sheet[cell_ref], bold=True)
                        logger.debug(f"Записано {value} в {sheet_name}!{cell_ref}")
            wb.save(str(excel_path))
            logger.info("Кол-во успешно записаны в АП")
        except Exception as e:
            logger.error(f"Ошибка при заполнении количеств: {e}")
            raise
        finally:
            if wb:
                wb.close()

    def fill_ap(self, excel_path: Path, photo_root: Path) -> None:
        """Заполняет АП данными о фотографиях.
        
        Args:
            excel_path: Путь к файлу Excel.
            photo_root: Корневая папка с фотографиями.
        """
        wb = None
        try:
            wb = load_workbook(str(excel_path), keep_vba=True)
            for sheet_name, (folder_name, col_name, col_folder) in self.FOLDER_MAP.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(f"Лист {sheet_name} не найден")
                    continue
                    
                sheet = wb[sheet_name]
                folder_path = photo_root / folder_name
                files_info = self.get_all_files_with_subfolders(folder_path)
                
                for i, file_info in enumerate(files_info):
                    r = self.ROW_START + i
                    sheet[f"{col_name}{r}"] = self.clean_name(file_info["filename"])
                    sheet[f"{col_folder}{r}"] = file_info["subfolder"]
                    self._style_cell(sheet[f"{col_name}{r}"])
                    self._style_cell(sheet[f"{col_folder}{r}"])
                    
                logger.debug(f"Обработано {len(files_info)} файлов для {sheet_name}")
                
            wb.save(str(excel_path))
            logger.info("АП успешно заполнен данными")
        except Exception as e:
            logger.error(f"Ошибка при заполнении АП: {e}")
            raise
        finally:
            if wb:
                wb.close()
